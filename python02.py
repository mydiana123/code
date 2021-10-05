import requests
import re
from lxml import etree
from rich.table import Table
from rich.console import Console

# 解析网址
url = 'https://terraria.fandom.com/zh/wiki/Boss'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML,like Gecko)'}
text = requests.get(url).text

# 解析boss名字
tree = etree.HTML(text)
name_list = tree.xpath('//li[@class="toclevel-2"]//span[@class="toctext"]/text()')

# 解析hp
regExp = '<[^>]*>'
new_text = re.sub(regExp, '', text)
# with open('./zhqe.txt', 'w', encoding='utf-8') as fp:
#     fp.write(new_text)
regExp2 = r'\d+,\d+ / \d+,\d+ / \d+,\d+ '
hp_list = re.findall(regExp2, new_text)
# print(hp_list)
newList = list(zip(name_list, hp_list))
# print(newList)


# 终端显示
console = Console()
console.print(':angry_face_with_horns:' + "[bold blue]this is terraria wiki" + ':angry_face_with_horns:')
table = Table(show_header=True, header_style='bold blue')
table.add_column('Name', style='dim', width=18)
table.add_column('hp', style='blink', width=30)
for items in newList:
    res1 = items[0]
    res2 = items[1]
    table.add_row(
        f"[bold magenta]{res1}",
        f"[green]{res2}",
    )
console.print(table)

# 打印表格
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
i, j = 1, 1
for name in name_list:
    ws['A' + str(i)] = name
    i += 1

for hp in hp_list:
    ws['B' + str(j)] = hp
    j += 1
    ws.merge_cells(f'B{str(j)}:D{j}')
wb.save('./terraria wiki.xlsx')
